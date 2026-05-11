"""
Export analysis results to Excel (.xlsx) and CSV formats.
"""
import csv
import io
from datetime import datetime
from urllib.parse import urlparse


def _safe_name(url: str) -> str:
    host = urlparse(url).netloc or "site"
    return host.replace(".", "_").replace(":", "_")


def export_csv(url: str, results: dict, output_path: str = None) -> str:
    """Write results to a CSV file. Returns the output path."""
    if not output_path:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = f"report_{_safe_name(url)}_{ts}.csv"

    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["URL", url])
        writer.writerow(["Generated", datetime.now().isoformat()])
        writer.writerow([])

        # Scores summary
        writer.writerow(["=== SCORES ==="])
        writer.writerow(["Section", "Score"])
        for sec in ("performance", "seo", "accessibility", "security"):
            s = results.get(sec, {}).get("score")
            writer.writerow([sec.title(), s if s is not None else "N/A"])
        writer.writerow([])

        # Issues
        writer.writerow(["=== ISSUES ==="])
        writer.writerow(["Section", "Severity", "Message"])
        for sec, data in results.items():
            if not isinstance(data, dict):
                continue
            for issue in data.get("issues", []):
                writer.writerow([
                    sec.title(),
                    issue.get("severity", "info"),
                    issue.get("msg", ""),
                ])

        # Performance metrics
        perf = results.get("performance", {})
        if perf.get("metrics"):
            writer.writerow([])
            writer.writerow(["=== CORE WEB VITALS ==="])
            writer.writerow(["Metric", "Value", "Status"])
            for label, m in perf["metrics"].items():
                writer.writerow([label, m.get("value", "N/A"), m.get("severity", "info")])

        # Links breakdown
        links = results.get("links", {}).get("details", {})
        if links:
            writer.writerow([])
            writer.writerow(["=== LINKS ==="])
            writer.writerow(["Total", links.get("total_links", 0)])
            writer.writerow(["Internal", links.get("internal", 0)])
            writer.writerow(["External", links.get("external", 0)])
            broken = links.get("broken", [])
            if broken:
                writer.writerow([])
                writer.writerow(["Broken Links"])
                writer.writerow(["Status", "URL"])
                for b in broken:
                    writer.writerow([b.get("status", "?"), b.get("url", "")])

    return output_path


def export_xlsx(url: str, results: dict, output_path: str = None) -> str:
    """Write results to an Excel file. Returns the output path."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl is required. Install with: pip install openpyxl")

    if not output_path:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = f"report_{_safe_name(url)}_{ts}.xlsx"

    wb = Workbook()

    # Style helpers
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    title_font = Font(bold=True, size=16, color="1E3A8A")
    sev_fills = {
        "critical": PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
        "warning": PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
        "ok": PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"),
        "info": PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid"),
    }
    thin = Side(border_style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ---- Summary sheet ----
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Web Analysis Report"
    ws["A1"].font = title_font
    ws.merge_cells("A1:D1")
    ws["A2"] = "URL:"
    ws["A2"].font = Font(bold=True)
    ws["B2"] = url
    ws["A3"] = "Generated:"
    ws["A3"].font = Font(bold=True)
    ws["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    ws["A5"] = "Section"
    ws["B5"] = "Score"
    ws["C5"] = "Status"
    for col in ("A5", "B5", "C5"):
        ws[col].font = header_font
        ws[col].fill = header_fill
        ws[col].alignment = Alignment(horizontal="center")
        ws[col].border = border

    row = 6
    for sec in ("performance", "seo", "accessibility", "security"):
        s = results.get(sec, {}).get("score")
        ws.cell(row=row, column=1, value=sec.title()).border = border
        ws.cell(row=row, column=2, value=s if s is not None else "N/A").border = border
        if s is not None:
            if s >= 90:
                status = "Good"
                fill = sev_fills["ok"]
            elif s >= 50:
                status = "Needs Work"
                fill = sev_fills["warning"]
            else:
                status = "Poor"
                fill = sev_fills["critical"]
        else:
            status = "N/A"
            fill = sev_fills["info"]
        cell = ws.cell(row=row, column=3, value=status)
        cell.fill = fill
        cell.border = border
        row += 1

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 18

    # ---- Issues sheet ----
    issues_ws = wb.create_sheet("Issues")
    headers = ["Section", "Severity", "Message"]
    for i, h in enumerate(headers, 1):
        c = issues_ws.cell(row=1, column=i, value=h)
        c.font = header_font
        c.fill = header_fill
        c.border = border
        c.alignment = Alignment(horizontal="center")
    row = 2
    for sec, data in results.items():
        if not isinstance(data, dict):
            continue
        for issue in data.get("issues", []):
            sev = issue.get("severity", "info")
            issues_ws.cell(row=row, column=1, value=sec.title()).border = border
            sev_cell = issues_ws.cell(row=row, column=2, value=sev.upper())
            sev_cell.fill = sev_fills.get(sev, sev_fills["info"])
            sev_cell.border = border
            sev_cell.font = Font(bold=True)
            msg_cell = issues_ws.cell(row=row, column=3, value=issue.get("msg", ""))
            msg_cell.border = border
            msg_cell.alignment = Alignment(wrap_text=True, vertical="top")
            row += 1
    issues_ws.column_dimensions["A"].width = 18
    issues_ws.column_dimensions["B"].width = 12
    issues_ws.column_dimensions["C"].width = 100

    # ---- Performance sheet ----
    perf = results.get("performance", {})
    if perf.get("metrics"):
        pws = wb.create_sheet("Performance")
        pws["A1"] = "Core Web Vitals"
        pws["A1"].font = title_font
        pws.merge_cells("A1:C1")
        for i, h in enumerate(["Metric", "Value", "Status"], 1):
            c = pws.cell(row=3, column=i, value=h)
            c.font = header_font
            c.fill = header_fill
            c.border = border
        row = 4
        for label, m in perf["metrics"].items():
            pws.cell(row=row, column=1, value=label).border = border
            pws.cell(row=row, column=2, value=m.get("value", "N/A")).border = border
            sev = m.get("severity", "info")
            sc = pws.cell(row=row, column=3, value=sev.upper())
            sc.fill = sev_fills.get(sev, sev_fills["info"])
            sc.border = border
            row += 1
        if perf.get("opportunities"):
            row += 1
            pws.cell(row=row, column=1, value="Opportunities").font = title_font
            row += 1
            for i, h in enumerate(["Title", "Savings (ms)"], 1):
                c = pws.cell(row=row, column=i, value=h)
                c.font = header_font
                c.fill = header_fill
                c.border = border
            row += 1
            for opp in sorted(perf["opportunities"], key=lambda x: -x.get("savings_ms", 0)):
                pws.cell(row=row, column=1, value=opp.get("title", "")).border = border
                pws.cell(row=row, column=2, value=opp.get("savings_ms", 0)).border = border
                row += 1
        pws.column_dimensions["A"].width = 40
        pws.column_dimensions["B"].width = 20
        pws.column_dimensions["C"].width = 12

    # ---- Links sheet ----
    links = results.get("links", {}).get("details", {})
    if links:
        lws = wb.create_sheet("Links")
        lws["A1"] = "Link Analysis"
        lws["A1"].font = title_font
        lws.merge_cells("A1:B1")
        lws["A3"] = "Total Links"
        lws["B3"] = links.get("total_links", 0)
        lws["A4"] = "Internal"
        lws["B4"] = links.get("internal", 0)
        lws["A5"] = "External"
        lws["B5"] = links.get("external", 0)
        for cell in ("A3", "A4", "A5"):
            lws[cell].font = Font(bold=True)
        broken = links.get("broken", [])
        if broken:
            lws["A7"] = "Broken Links"
            lws["A7"].font = title_font
            for i, h in enumerate(["Status", "URL"], 1):
                c = lws.cell(row=8, column=i, value=h)
                c.font = header_font
                c.fill = header_fill
                c.border = border
            for r, b in enumerate(broken, start=9):
                lws.cell(row=r, column=1, value=str(b.get("status", "?"))).border = border
                lws.cell(row=r, column=2, value=b.get("url", "")).border = border
        lws.column_dimensions["A"].width = 18
        lws.column_dimensions["B"].width = 80

    wb.save(output_path)
    return output_path
